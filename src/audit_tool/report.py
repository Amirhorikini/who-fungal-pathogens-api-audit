"""Consolidation of audit results into CSV (and Excel, if openpyxl is available)."""
import csv
from .common import log

FIELDNAMES = [
    "source", "pathogen_id", "pathogen_label",
    "api_score", "api_note",
    "documentacao_score", "documentacao_note",
    "integracao_score", "integracao_note",
    "http_status", "latency_ms", "raw_value", "probe_error", "probe_note",
]


def write_csv(rows, output_path):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    log(f"CSV report written -> {output_path}")


def write_excel(rows, output_path):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("openpyxl not installed -- skipped Excel export (CSV was already written). "
            "Install with 'pip install openpyxl' if you also want that format.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit"
    ws.append(FIELDNAMES)
    for row in rows:
        ws.append([row.get(f, "") for f in FIELDNAMES])

    for i, _ in enumerate(FIELDNAMES, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    wb.save(output_path)
    log(f"Excel report written -> {output_path}")
